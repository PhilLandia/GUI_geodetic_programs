import sys
import random

from PyQt5.QtWidgets import QApplication, QWidget, QComboBox, QSpinBox, QVBoxLayout, QPushButton, QMainWindow, \
    QLabel, QFormLayout, QDoubleSpinBox, QFrame, QGroupBox, QHBoxLayout, QFileDialog, QTableView, QDialog, QLineEdit, \
    QTableWidgetItem, QTableWidget, QMessageBox, QHeaderView, QScrollArea
from PyQt5.QtSql import QSqlDatabase, QSqlTableModel, QSqlQuery
from class_calculations import SoilLoadCalculator
from PyQt5.QtGui import QFont, QIntValidator, QRegularExpressionValidator
from PyQt5.QtCore import Qt, QRegularExpression
from well_layer_sampling import WellLayerSampling
import openpyxl
import xlwt

data = {
    'песок гравелистый плотный': [24, 50, 3, 8.075],
    'песок гравелистый ср': [16, 25, 3, 8.075],
    'песок гравелистый рыхлый': [10, 16, 3, 8.075],
    'песок крупный плотный': [16, 22, 3, 9.69],
    'песок крупный ср': [7, 18, 3, 9.69],
    'песок крупный рыхлый': [1, 5, 3, 9.69],
    'песок средний плотный': [10, 20, 4, 11.305],
    'песок средний ср': [6, 15, 5, 11.305],
    'песок средний рыхлый': [2.5, 5, 6, 11.305],
    'песок мелкий плотный': [5, 12, 7, 12.92],
    'песок мелкий ср': [4, 10, 7.5, 12.92],
    'песок мелкий рыхлый': [1.1, 4, 8, 12.92],
    'песок пылеватый плотный': [10, 12, 15, 20.1875],
    'песок пылеватый ср': [3, 10, 25.5, 24.225],
    'песок пылеватый рыхлый': [1, 3, 30, 29.07],
    'суглинок моренный твердый': [4, 6, 10, 40.375],
    'суглинок моренный полутвердый': [2.5, 4.7, 15, 48.45],
    'суглинок моренный тугопластичный': [2.25, 4.2, 20, 56.525],
    'суглинок моренный мягкопластичный': [1.375, 3.8, 25, 64.6],
    'суглинок моренный текучепластичный': [0.5, 0.8, 30, 72.675],
    'суглинки тяжелые, моренные полутвердые и тугопластичные': [3, 4.9, 10, 62.7],
    'суглинки тяжелые, моренные мягкопластичные': [1, 1.5, 15, 42.75],
    'пылеватые, аллювиальные, слабозаторфованые': [0.8, 1.5, 22, 52],
    'глины полутвердые и твердые': [2.5, 3.5, 50, 80],
    'глины твердые': [4, 6.5, 20, 40],
    'глины мягкопластичные': [1.5, 3.75, 30, 60],
    'суглинки юрские пластичные': [2.5, 3.5, 30, 50],
    'суглинок покровный мягкопластичный': [0.7, 2.9, 50, 90],
    'суглинок флювиогляциальный тугопластичный': [1, 3.3, 20, 60],
    'суглинок техногенн': [0.7, 5.5, 20, 56.525],
    'супесь пастичная сильнопесчанистая': [0.7, 5.5, 10, 40]
}


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.open_layers_select = None
        self.db = QSqlDatabase.addDatabase("QSQLITE")
        self.db.setDatabaseName("data1.db")
        self.layer_window = None
        self.data_from_db = None
        self.name_well = None

        self.setWindowTitle("Исследование скважин")
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)

        self.layout = QVBoxLayout()
        self.central_widget.setLayout(self.layout)
        self.layout.addWidget(QLabel("Название объекта"))
        self.name_well_edit = QLineEdit()
        self.layout.addWidget(self.name_well_edit)

        self.num_wells = QSpinBox()
        self.num_wells.setMinimum(1)
        self.num_wells.setMaximum(100)
        self.layout.addWidget(QLabel("Сколько скважин всего?"))
        self.layout.addWidget(self.num_wells)

        self.num_layers = QSpinBox()
        self.num_layers.setMinimum(1)
        self.num_layers.setMaximum(100)
        self.layout.addWidget(QLabel("Количество используемых слоев?"))
        self.layout.addWidget(self.num_layers)

        self.button_layers = QPushButton("Выбор слоев")
        self.button_layers.clicked.connect(self.open_layers_selection)
        self.layout.addWidget(self.button_layers)

        self.db_button = QPushButton("Редактирование таблицы")
        self.db_button.clicked.connect(self.open_edit_table_dialog)
        self.layout.addWidget(self.db_button)

        self.well_values = []
        self.well_fields = []

        self.data_from_db = self.fetch_data_from_db()

    def open_edit_table_dialog(self):
        edit_table_dialog = EditTableDialog(parent=self)
        edit_table_dialog.exec_()

    def open_layers_selection(self):
        self.name_well = self.name_well_edit.text()
        if self.name_well == '':
            self.name_well = 'Объект'
        self.open_layers_select = TableWidgetForLayer(self.num_layers.value(), self.data_from_db,
                                                      self.num_wells.value(), self.name_well)
        self.open_layers_select.show()
        self.hide()

    def fetch_data_from_db(self):
        db = QSqlDatabase.addDatabase("QSQLITE")
        db.setDatabaseName("data1.db")
        query = QSqlQuery(db)
        if db.open():
            if query.exec_("SELECT * FROM soil_data_new"):
                while query.next():
                    soil_type = query.value(0)
                    values = [query.value(i) for i in range(1, query.record().count())]
                    data[soil_type] = values
            else:
                print("Ошибка выполнения запроса:", query.lastError().text())

        return data


class TableWidgetForLayer(QWidget):
    def __init__(self, row_count, data_dict, num_wells, name_obj):
        super().__init__()
        self.window_layer_info = None
        self.num_wells = num_wells
        self.submit_button = None
        self.table_widget = None
        self.index_to_name = {}
        self.name_obj = name_obj
        self.row_count = row_count
        self.data_dict = data_dict
        self.combo_boxes = {}
        self.sample_counts = {}  # Словарь для хранения количества проб
        self.resize(820, 250)
        self.init_ui()

    def init_ui(self):
        self.table_widget = QTableWidget(self.row_count, 6)  # Увеличили количество столбцов на 1
        self.table_widget.setHorizontalHeaderLabels(
            ["Название слоя", "Количество проб", "Лобовое от", "Лобовое до", "Боковое от", "Боковое до"])

        for row in range(self.row_count):
            combo_box = QComboBox()
            combo_box.addItems(list(self.data_dict.keys()))
            combo_box.currentTextChanged.connect(lambda text, row=row: self.combo_box_changed(text, row))

            self.combo_boxes[row] = combo_box
            self.table_widget.setCellWidget(row, 0, combo_box)

            sample_count_item = QTableWidgetItem("10")  # Значение по умолчанию для количества проб
            self.table_widget.setItem(row, 1, sample_count_item)
            self.sample_counts[row] = 10  # Сохраняем начальное значение количества проб

            self.submit_button = QPushButton("Далее")
            self.submit_button.setEnabled(False)
            self.submit_button.clicked.connect(self.submit_button_clicked)

        layout = QVBoxLayout()
        layout.addWidget(self.table_widget)
        self.setLayout(layout)
        layout.addWidget(self.submit_button)

        self.table_widget.resizeColumnsToContents()

    def combo_box_changed(self, text, row):
        selected_item = self.data_dict.get(text)
        for col, value in enumerate(selected_item, start=2):  # Начинаем с 2-го столбца
            item = QTableWidgetItem(str(value))
            self.table_widget.setItem(row, col, item)
        self.check_data_completion()

    def check_data_completion(self):
        for row in range(self.row_count):
            for col in range(2, 6):  # Проверка начинается с 2-го столбца
                if self.table_widget.item(row, col) is None:
                    self.submit_button.setEnabled(False)
                    return
        self.submit_button.setEnabled(True)

    def submit_button_clicked(self):
        updated_data_dict, name_index, sample_counts = self.get_updated_data_dict()
        self.window_layer_info = WellInputWindow(self.num_wells, updated_data_dict, self.name_obj, name_index,
                                                 sample_counts)
        self.window_layer_info.show()
        self.hide()
        print("Updated Data:", updated_data_dict)
        print("Sample Counts:", sample_counts)  # Выводим количество проб

    def get_updated_data_dict(self):
        updated_dict = {}
        name_index = {}
        sample_counts = {}
        for row, combo_box in self.combo_boxes.items():
            current_name = combo_box.currentText()

            row_data = []
            for col in range(2, 6):
                item = self.table_widget.item(row, col)
                if col in [2, 3]:
                    value = float(item.text())
                    random_percentage = random.uniform(-0.1, 0.1)
                    random_value = value * (1 + random_percentage)
                    row_data.append(random_value)
                else:
                    row_data.append(float(item.text()))
            row_data.append(current_name)
            updated_dict[row + 1] = row_data
            sample_count_item = self.table_widget.item(row, 1)
            sample_counts[row + 1] = int(sample_count_item.text())

            if current_name in name_index:
                name_index[current_name] += f",{row + 1}"  # Добавляем номер строки
            else:
                name_index[current_name] = str(row + 1)  # Создаем запись с номером строки

        return updated_dict, name_index, sample_counts


class EditTableDialog(QDialog):
    def __init__(self, parent=MainWindow):
        super().__init__(parent)
        self.setWindowTitle("Редактирование таблицы")
        self.db = QSqlDatabase.database()
        layout = QVBoxLayout(self)

        self.table_view = QTableView(self)
        layout.addWidget(self.table_view)

        self.model = QSqlTableModel(self)
        self.model.setTable('soil_data_new')
        self.model.select()
        self.table_view.setModel(self.model)
        # Added these two lines
        self.table_view.resizeColumnsToContents()
        table_width = self.table_view.horizontalHeader().length() + self.table_view.verticalHeader().width() + 100
        table_height = self.table_view.verticalHeader().length() + self.table_view.horizontalHeader().height()

        # Установите размер диалогового окна в соответствии с размерами таблицы
        self.resize(table_width, table_height)

        # ... ваш код ...

        # Установите горизонтальное растяжение для о

        button_layout = QVBoxLayout()
        save_button = QPushButton("Сохранить")
        save_button.clicked.connect(self.save_changes)
        button_layout.addWidget(save_button)

        delete_button = QPushButton("Удалить")
        delete_button.clicked.connect(self.delete_row)
        button_layout.addWidget(delete_button)

        add_button = QPushButton("Добавить")
        add_button.clicked.connect(self.add_row)
        button_layout.addWidget(add_button)

        exit_button = QPushButton("Выход")
        exit_button.clicked.connect(self.accept)
        button_layout.addWidget(exit_button)

        layout.addLayout(button_layout)

        # Создание и настройка соединения с базой данных

    def save_changes(self):
        self.model.submitAll()

    def delete_row(self):
        selected_rows = self.table_view.selectionModel().selectedRows()
        for row in selected_rows:
            self.model.removeRow(row.row())
        self.model.submitAll()

    def add_row(self):
        self.model.insertRow(self.model.rowCount())

    def closeEvent(self, event):
        self.save_changes()
        self.db.close()


class WellInputWindow(QMainWindow):
    def __init__(self, num_wells, layers_db, name_obj, name_index, sample_counts):
        super().__init__()
        self.well_blocks = None
        self.used_layers = None
        self.name_index = None
        self.name_obj = None
        self.sample_counts = None
        self.depth_window = None
        self.data = None
        self.dataframes = None
        self.initUI(num_wells, layers_db, name_obj, name_index, sample_counts)

    def initUI(self, num_wells, layers_db, name_obj, name_index, sample_counts):
        
        self.depth_window = None
        self.dataframes = None

        self.data = {}
        
        self.sample_counts = sample_counts
        self.name_obj = name_obj
        
        self.name_index = name_index
        print(name_index)
        
        self.used_layers = layers_db
        self.setWindowTitle("Input Wells")
        self.setGeometry(100, 100, 500, 800)

        # Create a scroll area
        scroll_area = QScrollArea(self)
        scroll_area.setWidgetResizable(True)

        # Create a widget to hold your main layout
        main_widget = QWidget()
        scroll_area.setWidget(main_widget)

        self.setCentralWidget(scroll_area)

        main_layout = QVBoxLayout()  # Main vertical layout
        main_widget.setLayout(main_layout)
        
        self.well_blocks = []

        wells_layout = QHBoxLayout()  # Horizontal layout for well blocks

        for i in range(num_wells):
            well_block = self.create_well_block(i + 1)
            self.well_blocks.append(well_block)
            wells_layout.addLayout(well_block)

        main_layout.addLayout(wells_layout)  # Add horizontal layout to the main layout

        finish_button = QPushButton("Завершить редактирование")
        finish_button.clicked.connect(self.finish_editing)
        main_layout.addWidget(finish_button, alignment=Qt.AlignCenter)
        for value, key in self.name_index.items():
            label = QLabel(f"{key} - {value}")
            font = QFont()  # Create a font object
            font.setPointSize(14)  # Set font size (in this case, 16)
            label.setFont(font)
            main_layout.addWidget(label)

    def create_well_block(self, well_number):
        onlyInt = QIntValidator()
        onlyInt.setRange(0, 500)
        regExp = QRegularExpression("^[1-9][0-9]*(?:[.,][0-9]*)?$")

        validator_float = QRegularExpressionValidator(regExp)
        well_block_layout = QVBoxLayout()

        well_label = QLabel(f"Скважина {well_number}")
        well_block_layout.addWidget(well_label)

        well_table = QTableWidget()
        well_table.setRowCount(12)
        well_table.setColumnCount(2)
        well_table.setHorizontalHeaderLabels(["Название слоя", "Глубина подошвы"])
        self.data[well_number] = (well_table, [])  # Initialize list for well data
        for i in range(12):
            line_edit = QLineEdit()
            line_edit.setValidator(onlyInt)
            depth_line_edit = QLineEdit()
            depth_line_edit.setValidator(validator_float)
            well_table.setCellWidget(i, 0, line_edit)
            well_table.setCellWidget(i, 1, depth_line_edit)
        well_table.setColumnWidth(0, 10)
        well_table.setColumnWidth(1, 10)

        well_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        well_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)

        well_table.horizontalHeader().setMinimumSectionSize(10)
        well_table.resizeColumnsToContents()
        well_block_layout.addWidget(well_table)

        ugv_label = QLabel("Значение УГВ:")
        well_block_layout.addWidget(ugv_label)

        ugv_line = QLineEdit()
        ugv_line.setValidator(validator_float)
        well_block_layout.addWidget(ugv_line)

        label_name_well = QLabel("Название скважины:")
        well_block_layout.addWidget(label_name_well)

        well_name = QLineEdit()
        well_block_layout.addWidget(well_name)

        self.data[well_number] = (well_table, ugv_line, well_name)
        return well_block_layout

    def keyPressEvent(self, event):
        if event.key() == 16777220:  # Код клавиши Enter
            self.finish_editing()

    def create_excel_table(self, samples):
        unique_elements = {}
        for key, values in samples.items():
            for value, weight in values:
                unique_elements[value] = weight

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Добавляем названия столбцов
        sheet.cell(row=1, column=1, value="Номер скважины")
        sheet.cell(row=1, column=2, value="Номер слоя")
        sheet.cell(row=1, column=3, value="Глубина подошвы")

        row_num = 2  # Начинаем со второй строки после названий столбцов

        for key, values in samples.items():
            for value, weight in values:
                sheet.cell(row=row_num, column=1, value=key)
                sheet.cell(row=row_num, column=2, value=value)
                sheet.cell(row=row_num, column=3, value=weight)
                row_num += 1

        workbook.save(f'output_{self.name_obj}.xlsx')

    def finish_editing(self):
        well_data, list_ugv, well_names = self.get_data()
        if well_data is None or list_ugv is None or well_names is None:
            return
        self.create_excel_table(well_data)
        soil_calculator = SoilLoadCalculator()
        print(self.used_layers)
        for borehole_num in range(len(well_data)):
            soil_calculator.generate_data(borehole_num + 1, self.used_layers, well_data)
            soil_calculator.process_data(list_ugv, borehole_num)
        dataframes = soil_calculator.df_list
        self.depth_window = DepthRangeWindow(dataframes, self.name_obj, well_names, self.sample_counts)
        self.depth_window.show()
        self.hide()
        # print("Well Data:", well_data)
        # print("Spinbox:", list_spin)
        # print("Well Names:", well_names)

    def get_data(self):
        samples = {}
        well_names = []
        ugv_list = []
        all_indexes = []
        num_layers = {}
        for indexes in self.name_index.values():
            index_list = indexes.split(',')
            all_indexes.extend(index_list)

        all_indexes = list(set(map(int, all_indexes)))  # Преобразуем в целые числа и удаляем дубликаты
        all_indexes.sort()
        error_message = None  # Сохраняем сообщение об ошибке, если есть

        for well_number, (well_table, line_ugv, well_name) in self.data.items():
            well_data = []

            well_names.append(well_name.text() or f'Скважина_{well_number}')

            if line_ugv.text() == '':
                ugv_value = 1000
            else:
                ugv_value = float(line_ugv.text().replace(',', '.'))
            ugv_list.append(ugv_value)  # Получаем название скважины
            prev_depth = None  # Предыдущая глубина

            for row in range(well_table.rowCount()):
                line_edit = well_table.cellWidget(row, 0)
                depth_line_edit = well_table.cellWidget(row, 1)
                if depth_line_edit.text() is None or depth_line_edit.text() == '':
                    break
                depth_value = float(depth_line_edit.text().replace(',', '.'))
                selected_item = int(line_edit.text())

                if selected_item is None or selected_item == '' or (
                        prev_depth is not None and depth_value < prev_depth):
                    break
                if selected_item not in all_indexes:
                    error_message = "Вводите только числа из словаря"
                    break
                if selected_item not in num_layers.keys():
                    num_layers[selected_item] = 1
                else:
                    num_layers[selected_item] += 1
                if num_layers[selected_item] > self.sample_counts[selected_item]:
                    error_message = "Слой используется чаще чем пробы"
                    break
                well_data.append((selected_item, depth_value))
                prev_depth = depth_value

            if error_message:
                break
            else:
                samples[well_number] = well_data

        if error_message:
            QMessageBox.warning(self, "Предупреждение", error_message)
            return None, None, None

        return samples, ugv_list, well_names


class DepthRangeWindow(QMainWindow):
    def __init__(self, dataframes, name_obj, wells_names, sample):
        super().__init__()
        self.last_window = None
        self.folder_path = ''
        self.dataframes = dataframes
        self.range_start_spinboxes = []
        self.range_end_spinboxes = []
        self.name_obj = name_obj
        self.wells_names = wells_names
        self.sample_number = sample
        self.setWindowTitle("Выбор диапазона глубины")

        central_widget = QWidget(self)
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        for i, dataframe in enumerate(dataframes):
            well_label = QLabel(f"Скважина: {i + 1}", self)
            layout.addWidget(well_label)

            start_spinbox = QDoubleSpinBox(self)
            end_spinbox = QDoubleSpinBox(self)
            start_spinbox.setDecimals(1)
            start_spinbox.setValue(0.4)
            max_value = dataframe.iloc[-1]['Глубина']
            end_spinbox.setDecimals(1)
            end_spinbox.setValue(max_value)
            start_row = QWidget()
            start_layout = QHBoxLayout(start_row)
            start_layout.addWidget(QLabel("Начало:"))
            start_layout.addWidget(start_spinbox)
            layout.addWidget(start_row)

            end_row = QWidget()
            end_layout = QHBoxLayout(end_row)
            end_layout.addWidget(QLabel("Конец:"))
            end_layout.addWidget(end_spinbox)
            layout.addWidget(end_row)

            self.range_start_spinboxes.append(start_spinbox)
            self.range_end_spinboxes.append(end_spinbox)

        load_button = QPushButton("Загрузить данные", self)
        layout.addWidget(load_button)
        load_button.clicked.connect(self.load_data)

        folder_button = QPushButton("Выбрать папку", self)
        layout.addWidget(folder_button)
        folder_button.clicked.connect(self.select_folder)

    def load_data(self):
        for i, df in enumerate(self.dataframes):
            range_start = self.range_start_spinboxes[i].value()
            range_end = self.range_end_spinboxes[i].value()

            filtered_data = df[
                (df["Глубина"] >= range_start) &
                (df["Глубина"] <= range_end)
                ]
            column_mapping = {
                'Глубина': 'Глубина',
                'Округленное_Среднее_Гаусса': 'СР1',
                'Округленный_Средний_Нов.Бок.Лоб': 'СР2'
            }
            if self.folder_path == '':
                self.save_to_xls(filtered_data, f"{self.name_obj}_ТСЗ_{self.wells_names[i]}.xls", column_mapping)
            else:
                self.save_to_xls(filtered_data, f"{self.folder_path}/{self.name_obj}_{self.wells_names[i]}.xls",
                                 column_mapping)  # Сохраняем
            # отфильтрованный датафрейм

    def save_to_xls(self, df, filename, column_mapping=None):
        if column_mapping is None:
            column_mapping = df.columns.to_dict()

        workbook = xlwt.Workbook()
        worksheet = workbook.add_sheet('Sheet1')

        # Записываем данные
        for row_idx, (_, row) in enumerate(df.iterrows()):
            for col_idx, (old_name, _) in enumerate(column_mapping.items(), start=1):
                worksheet.write(row_idx, col_idx, row[old_name])

        workbook.save(filename)
        self.last_window = LastWindow(self.sample_number, self.name_obj, self.folder_path)
        self.last_window.show()

    def select_folder(self):
        options = QFileDialog.Options()
        self.folder_path = QFileDialog.getExistingDirectory(self, "Выберите папку", options=options)
        if self.folder_path:
            print("Выбрана папка:", self.folder_path)
        else:
            self.folder_path = None


class LastWindow(QWidget):
    def __init__(self, num_samples, name_obj, folder_path):
        super().__init__()
        self.num_samples = num_samples
        self.name = name_obj
        self.sampling = None
        self.label = None
        self.folder_path = folder_path
        self.line_edit = None
        self.button = None
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Номер начальной пробы")

        layout = QVBoxLayout()

        self.label = QLabel("Введите номер начальной пробы")
        layout.addWidget(self.label)

        self.line_edit = QLineEdit()
        layout.addWidget(self.line_edit)

        self.button = QPushButton("Завершить")
        self.button.clicked.connect(self.button_clicked)
        layout.addWidget(self.button)

        self.setLayout(layout)

    def button_clicked(self):
        self.sampling = WellLayerSampling(f'output_{self.name}.xlsx')
        self.sampling.load_data()
        self.sampling.find_layer_boundaries()
        self.sampling.calculate_total_depths()
        number_samples = int(self.line_edit.text())
        self.sampling.num_samples = self.num_samples
        self.sampling.calculate_layer_depths()
        self.sampling.generate_samples(number_samples)
        if self.folder_path:
            self.sampling.save_samples_to_excel('template.xls', f'{self.folder_path}/Импорт_физики {self.name}.xls')
        else:
            self.sampling.save_samples_to_excel('template.xls', f'Импорт_физики {self.name}.xls')
        sys.exit(0)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_window = MainWindow()
    main_window.show()
    sys.exit(app.exec_())
